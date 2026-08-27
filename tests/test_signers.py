"""Подписанты из payload Doc-V (режим снимка) в реестрах."""
import io
import json
from pathlib import Path

import openpyxl

from conftest import docv_headers

MODEL = json.loads((Path(__file__).parent / "data" / "model.json").read_text(encoding="utf-8"))

SIGNERS = {
    "soglasovano": {"fio": "Аманов Б.Ш.", "position": "Генеральный директор",
                    "company": "ТОО «Шар-Кұрылыс»"},
    "utverzhdayu": {"fio": "Акшалов Т.М.", "position": "Директор",
                    "company": 'ТОО "СМУ Аргон"'},
    "coordinators": [
        {"fio": "Абдрахманова Х.М.", "position": "Главный бухгалтер",
         "company": "ТОО «Шар-Кұрылыс»"},
        {"fio": "Бекмуратов Е.И.", "position": "Начальник ЮО",
         "company": "", "mark": "СОГЛАСОВАНО"},
    ],
}


def _sheet(client, payload, kind="inner"):
    r = client.post(f"/render/registry/{kind}", json=payload, headers=docv_headers())
    assert r.status_code == 200, r.text
    token = r.json()["download_url"].rsplit("/", 1)[1]
    wb = openpyxl.load_workbook(io.BytesIO(client.get(f"/files/{token}").content))
    return wb[[s for s in wb.sheetnames if not s.startswith("СПР_")][0]]


def test_inner_snapshot_mode(client):
    entry = dict(MODEL["request"][0], signers=SIGNERS)
    sheet = _sheet(client, {"request": [entry]})
    # верхние блоки — значения, не формулы
    assert sheet["F2"].value == "СОГЛАСОВАНО:"
    assert sheet["F3"].value == "Генеральный директор ТОО «Шар-Кұрылыс»"
    assert sheet["F5"].value.endswith("Аманов Б.Ш.") and "_" in sheet["F5"].value
    assert sheet["I3"].value == 'Директор ТОО "СМУ Аргон"'
    assert sheet["I5"].value.endswith("Акшалов Т.М.")
    for ref in ("F2", "F3", "F5", "I3", "I5"):
        assert not str(sheet[ref].value).startswith("=")
    # согласующие значениями, без ID в колонке B
    body = [(r, sheet.cell(row=r, column=6).value, sheet.cell(row=r, column=9).value)
            for r in range(18, sheet.max_row + 1)
            if sheet.cell(row=r, column=6).value or sheet.cell(row=r, column=9).value]
    texts = [str(v6) + "|" + str(v9) for _, v6, v9 in body]
    assert any("Главный бухгалтер" in t and "Абдрахманова" in t for t in texts)
    # запись с mark: заголовок отдельной строкой, подпись двумя ниже
    mark_row = next(r for r, v6, _ in body if v6 == "СОГЛАСОВАНО")
    assert sheet.cell(row=mark_row + 2, column=9).value == "Бекмуратов Е.И."
    assert not any(isinstance(sheet.cell(row=r, column=2).value, int)
                   for r in range(18, sheet.max_row + 1))
    assert not any("VLOOKUP" in str(v6) for _, v6, _ in body)


def test_inner_without_utverzhdayu_hides_label(client):
    signers = {"soglasovano": SIGNERS["soglasovano"], "coordinators": []}
    entry = dict(MODEL["request"][0], signers=signers)
    sheet = _sheet(client, {"request": [entry]})
    assert sheet["I2"].value in ("", None)  # статичный «УТВЕРЖДАЮ» убран
    assert sheet["I5"].value in ("", None)


def test_inner_without_signers_keeps_yaml_mode(client):
    sheet = _sheet(client, {"request": [dict(MODEL["request"][0])]})
    assert str(sheet["F3"].value).startswith("=IF")  # формулы шаблона на месте


def test_priority_signers_override(client):
    entry = dict(MODEL["request"][0], signers={"coordinators": [
        {"position": "Финансовый директор", "fio": "Омарова Г.А."},
        {"position": "Начальник ПТО", "fio": "Королькова Е. В."},
        {"position": "Исполнительный директор", "fio": "Сергачев П.А."},
    ]})
    sheet = _sheet(client, {"request": [entry]}, kind="priority")
    joined = "|".join(str(c.value) for row in sheet.iter_rows(min_col=2, max_col=4)
                      for c in row if c.value)
    assert "Финансовый директор" in joined and "Омарова Г.А." in joined
    assert "Сергачев П.А." in joined
