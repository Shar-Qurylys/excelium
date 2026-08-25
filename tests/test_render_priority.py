import io
import json
from pathlib import Path

import openpyxl

from conftest import docv_headers

MODEL = json.loads((Path(__file__).parent / "data" / "model.json").read_text(encoding="utf-8"))


def test_contract(client):
    r = client.post("/render/registry/priority", json=MODEL, headers=docv_headers())
    assert r.status_code == 200, r.text
    token = r.json()["download_url"].rsplit("/", 1)[1]
    wb = openpyxl.load_workbook(io.BytesIO(client.get(f"/files/{token}").content))
    assert "REESTR" not in wb.sheetnames
    objects = {str(d.get("object_name") or "Без объекта") for d in MODEL["request"]}
    assert len(wb.sheetnames) == len(objects)
    sheet = wb["Администрация"]
    assert sheet["A4"].value == "Объект: Администрация"
    rows = [r for r in range(7, sheet.max_row + 1) if sheet.cell(row=r, column=1).value]
    n = sum(1 for d in MODEL["request"] if d["object_name"] == "Администрация")
    assert len(rows) == n
    total_row = 7 + n
    assert sheet[f"D{total_row}"].value == "ВСЕГО"
    assert abs(sheet[f"E{total_row}"].value -
               sum(float(d["payment_sum"]) for d in MODEL["request"]
                   if d["object_name"] == "Администрация")) < 1e-6
    assert sheet.print_area
