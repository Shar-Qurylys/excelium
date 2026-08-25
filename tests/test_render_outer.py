import io
import json
from pathlib import Path

import openpyxl

from conftest import docv_headers

MODEL = json.loads((Path(__file__).parent / "data" / "model.json").read_text(encoding="utf-8"))


def _render(client, payload):
    r = client.post("/render/registry/outer", json=payload, headers=docv_headers())
    assert r.status_code == 200, r.text
    token = r.json()["download_url"].rsplit("/", 1)[1]
    return openpyxl.load_workbook(io.BytesIO(client.get(f"/files/{token}").content))


def test_contract(client):
    wb = _render(client, MODEL)
    sheet = wb.active
    assert str(sheet["B7"].value).startswith("Разрешение № от ")
    assert "." in str(sheet["B7"].value)  # дата, а не сырой timestamp
    n = len(MODEL["request"])
    assert sheet["B13"].value == 1 and sheet[f"B{12 + n}"].value == n
    assert isinstance(sheet["G13"].value, (int, float)) and isinstance(sheet["H13"].value, (int, float))
    # реквизиты собраны с названием банка
    first = MODEL["request"][0]
    assert first["schet_counter"] in sheet["I13"].value
    # блок подписей после позиций
    values = [sheet.cell(row=r, column=2).value for r in range(13, sheet.max_row + 1)]
    assert any(v and "New Line Project" in str(v) for v in values)
    assert any(v == "Директор:" for v in values)


def test_unknown_bik_labeled_with_itself(client):
    payload = {"request": [dict(MODEL["request"][0], BIK_counter="XXXXKZKA")]}
    sheet = _render(client, payload).active
    assert "в XXXXKZKA" in sheet["I13"].value  # раньше был KeyError


def test_missing_requisites_leaves_cell_empty(client):
    payload = {"request": [dict(MODEL["request"][0], schet_counter="", BIK_counter="")]}
    sheet = _render(client, payload).active
    assert sheet["I13"].value is None


def test_no_payment_type_no_none_prefix(client):
    payload = {"request": [dict(MODEL["request"][0], payment_type="")]}
    sheet = _render(client, payload).active
    assert not str(sheet["E13"].value).startswith("None")  # раньше «None к договору …»
