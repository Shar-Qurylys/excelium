import io
import json
from pathlib import Path

import openpyxl

from conftest import docv_headers

MODEL = json.loads((Path(__file__).parent / "data" / "model.json").read_text(encoding="utf-8"))


def _render(client, payload=None):
    r = client.post("/render/registry/inner", json=payload or MODEL, headers=docv_headers())
    assert r.status_code == 200, r.text
    url = r.json()["download_url"]
    token = url.rsplit("/", 1)[1]
    resp = client.get(f"/files/{token}")
    assert resp.status_code == 200
    return openpyxl.load_workbook(io.BytesIO(resp.content))


def test_contract(client):
    wb = _render(client)
    # служебные листы скрыты, REESTR удалён
    assert "REESTR" not in wb.sheetnames
    assert wb["СПР_ПОДПИСАНТОВ"].sheet_state == "hidden"
    assert wb["СПР_ОБЪЕКТОВ"].sheet_state == "hidden"
    data_sheets = [s for s in wb.sheetnames if not s.startswith("СПР_")]
    # 13 позиций дают лист на каждую пару компания+объект
    pairs = {(d["organization"], d["object_name"]) for d in MODEL["request"]}
    assert len(data_sheets) == len(pairs)

    sheet = next(wb[s] for s in data_sheets if s.endswith("Администрация")
                 and wb[s]["F17"].value and "Шар-Кұрылыс" in wb[s]["F17"].value)
    assert sheet["G11"].value == "Администрация"
    assert str(sheet["F7"].value).startswith("РЕЕСТР ПЛАТЕЖЕЙ №20/")
    # позиция: стороны, сумма числом, назначение собрано
    assert sheet["F17"].value.startswith("Заявитель: ")
    assert isinstance(sheet["H17"].value, (int, float))
    assert sheet["I17"].value
    # подписанты: директора в B2/B4, ниже — ID с формулами
    assert sheet["B2"].value == 0 and sheet["B4"].value == 1
    ids = [sheet.cell(row=r, column=2).value for r in range(18, sheet.max_row + 1)
           if isinstance(sheet.cell(row=r, column=2).value, int)]
    assert ids, "нет строк подписантов"
    formulas = [sheet.cell(row=r, column=6).value for r in range(18, sheet.max_row + 1)
                if str(sheet.cell(row=r, column=6).value or "").startswith("=IFERROR")]
    assert "СПР_ПОДПИСАНТОВ" in formulas[0]
    assert sheet.print_area


def test_empty_zatraty_does_not_crash(client):
    payload = {"request": [dict(MODEL["request"][0], zatraty=None)]}
    wb = _render(client, payload)
    assert wb  # раньше падало на .lower() от None


def test_expense_type_excludes_approvers(client):
    base = dict(MODEL["request"][0], organization='ТОО "СМУ Аргон"',
                object_name='ЖК "New Line"')  # список list_3 содержит ID 6
    wb_normal = _render(client, {"request": [dict(base, zatraty="СМР")]})
    wb_salary = _render(client, {"request": [dict(base, zatraty="Зарплата")]})

    def ids(wb):
        sheet = wb[[s for s in wb.sheetnames if not s.startswith("СПР_")][0]]
        return [sheet.cell(row=r, column=2).value for r in range(18, sheet.max_row + 1)
                if isinstance(sheet.cell(row=r, column=2).value, int)]

    assert 6 in ids(wb_normal)
    assert 6 not in ids(wb_salary)


def test_unknown_company_yields_blank_signatures(client, caplog):
    payload = {"request": [dict(MODEL["request"][0], organization="ТОО «Никто»",
                                object_name="Нигде")]}
    wb = _render(client, payload)
    sheet = wb[[s for s in wb.sheetnames if not s.startswith("СПР_")][0]]
    assert sheet["B2"].value == 0 and sheet["B4"].value == 0
