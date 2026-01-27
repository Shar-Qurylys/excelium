from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
import openpyxl
import json
from datetime import datetime
import zipfile
import os
import re
from dataclasses import dataclass
from typing import Optional


def set_border(style):
    border = Border(left=Side(style=style),
                right=Side(style=style),
                top=Side(style=style),
                bottom=Side(style=style))
    return border

def format_row(sheet, row_number, columns):
    # Define the font, border, and alignment
    font = Font(name='Arial', size=12)

    #set thin border
    border = set_border('thin')

    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Create a fill object with yellow color
    yellow_fill = PatternFill(start_color='FFFF00',
                          end_color='FFFF00',
                          fill_type='solid')
    # Set the row height
    sheet.row_dimensions[row_number].height = 100

    # Apply formatting to each cell in the row
    for col in columns:
        cell = sheet[f'{col}{row_number}']
        cell.font = font
        cell.border = border
        cell.alignment = alignment
        sheet[f'F{row_number}'].alignment = Alignment(horizontal='left', wrapText=True)

def format_row_priority_registry(sheet, row_number, columns):
    # Define the font, border, and alignment
    font = Font(name='Arial', size=12)

    #set thin border
    border = set_border('thin')

    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Create a fill object with yellow color
    yellow_fill = PatternFill(start_color='FFFF00',
                          end_color='FFFF00',
                          fill_type='solid')
    # Set the row height
    sheet.row_dimensions[row_number].height = 75

    # Apply formatting to each cell in the row
    for col in columns:
        cell = sheet[f'{col}{row_number}']
        cell.font = font
        cell.border = border
        cell.alignment = alignment

def set_cell_properties(sheet, row, column, value, border=None, alignment=None, font=None):
    '''

    Needs to be integrated into inner registry
    Sets the properties of a cell in a given sheet.

    Parameters:
    - sheet: The sheet object where the cell is located.
    - row: The row index of the cell.
    - column: The column index of the cell.
    - value: The value to be set in the cell.
    - border: (optional) The border style to be applied to the cell.
    - alignment: (optional) The alignment style to be applied to the cell.
    - font: (optional) The font style to be applied to the cell.

    Returns:
    - None
    '''

    cell = sheet.cell(row=row, column=column, value=value)
    if border:
        cell.border = border
    if alignment:
        cell.alignment = alignment
    if font:
        cell.font = font

def hide_sheets(ab, ss):
    for s in ss:
        ab[s].sheet_state = 'hidden'

def find_last_row_in_col(sheet, col_index):
    """
    Find the last non-empty row in a specific column.

    :param sheet: The worksheet object.
    :param col_index: The index of the column to search, starting from 1 for column A.
    :return: The row number of the last non-empty cell, or None if the column is empty.
    """
    # Openpyxl is 1-indexed, but using max_row directly as start makes the code clearer
    for row in range(sheet.max_row, 0, -1):
        if sheet.cell(row=row, column=col_index).value:
            return row
    return None

def load_excel(f):
    ab1 = openpyxl.load_workbook(f)
    return ab1

def read_json():
    '''
    Returns sorted json file from the model data
    Component of the testing module
    '''

    with open('tests/model.json', 'r') as file:
        json_data = json.load(file)

        payment_documents = json_data.get('request', [])

        sorted_payment_documents = sorted(payment_documents, key=lambda x: x.get('object_name', ''))

        json_data = {'request': sorted_payment_documents}
        return json_data

def format_datetime(datetime_str):
    # Parse the datetime string
    try:
        dt = datetime.fromisoformat(datetime_str)

        # Format the datetime into "dd/mm/yyyy" format
        formatted_date = dt.strftime("%d/%m/%Y")
        return formatted_date
    except ValueError:
        return datetime_str


# ============================================================================
# Field Configuration for JSON Parsing
# ============================================================================

@dataclass
class FieldDef:
    """Definition for a JSON field to be included in concatenated info."""
    key: str                          # JSON key name
    prefix: str = ''                  # Text prefix to add
    suffix: str = ''                  # Text suffix to add  
    strip_prefix: str = ''            # Text to strip from the start
    placeholder: Optional[str] = None # Value to exclude (e.g., 'placeholder')
    extract_numbers: bool = False     # Use regex to extract numbers first

# Data-driven field configuration - easy to add/modify/remove fields
FIELD_CONFIG = [
    FieldDef(key='schet_na_oplatu', prefix='Счет на оплату №', strip_prefix='№'),
    FieldDef(key='esf', prefix='ЭСФ №', strip_prefix='№'),
    FieldDef(key='avr', prefix='Акт выполненных работ №', strip_prefix='№'),
    FieldDef(key='akt_sverki', prefix='Акт сверки ', strip_prefix='№'),
    FieldDef(key='sluzhebnaja_zapiska', prefix='Служебная записка '),
    FieldDef(key='avansovy_otchet', prefix='Авансовый отчет №', strip_prefix='№'),
    FieldDef(key='TRU'),
    FieldDef(key='letter', prefix='Письмо '),
    FieldDef(key='mediation', prefix='Медиация/Решение суда №', strip_prefix='№'),
    FieldDef(key='nakladnye', prefix='Накладные: ', extract_numbers=True),
    FieldDef(key='sogl_o_rastor', prefix='Согл. о расторжении №', strip_prefix='№', 
             placeholder='placeholder', extract_numbers=True),
    FieldDef(key='prilozhenija', prefix='по приложению ', strip_prefix='Приложение '),
]


def _process_field(data_item: dict, field: FieldDef) -> Optional[str]:
    """Process a single field according to its definition."""
    value = data_item.get(field.key, '')
    if not value:
        return None
    
    value = str(value).strip()
    
    # Skip placeholder values
    if field.placeholder and value == field.placeholder:
        return None
    
    # Extract numbers if required
    if field.extract_numbers:
        match = re.search(r'\d+', value)
        if not match:
            return None
        value = value[match.start():].strip()
    
    # Strip prefix if specified
    if field.strip_prefix:
        value = value.lstrip(field.strip_prefix)
    
    if not value:
        return None
    
    return f"{field.prefix}{value}{field.suffix}"


def create_concatenated_info(data_item: dict) -> str:
    """
    Creates a concatenated string of payment information from JSON data.
    
    Uses FIELD_CONFIG for data-driven field processing. Special handling
    for payment_type, payment_objective, contract info, and payment number.
    """
    parts = []
    
    # Special handling: payment_type and payment_objective
    payment_type = data_item.get('payment_type', '').strip()
    payment_objective = data_item.get('payment_objective', '').strip()
    
    # Add payment type only if different from objective
    if payment_type and payment_type != payment_objective:
        parts.append(payment_type)
    
    if payment_objective:
        parts.append(payment_objective)
    
    # Process all configured fields
    for field in FIELD_CONFIG:
        result = _process_field(data_item, field)
        if result:
            parts.append(result)
    
    # Special handling: zusaetzliches_vertrag / contract
    zusaetzliches_vertrag = data_item.get('zusaetzliches_vertrag', '')
    if zusaetzliches_vertrag and zusaetzliches_vertrag != 'placeholder':
        zv_text = str(zusaetzliches_vertrag).lstrip('Доп. соглашение')
        parts.append(f'ДС {zv_text}')
    else:
        name_of_contract = data_item.get('name_of_contract', '')
        date_of_contract = data_item.get('date_of_contract', '')
        if name_of_contract and date_of_contract:
            parts.append(f"Дог. {name_of_contract}")
    
    # Special handling: payment_number
    payment_number = data_item.get('payment_number', '')
    doctype = data_item.get('doctype', '')
    if payment_number:
        parts.append(f"{doctype} №{payment_number}")
    
    return ", ".join(parts).rstrip(", ")

def add_colontituls(sheet):
    sheet.oddFooter.left.text = "Группа компаний «Шар Құрылыс»" # Left footer
    sheet.oddFooter.right.text = "Дата и время печати &D &T" # Right footer

def set_print_area(sheet):
    last_row = find_last_row_in_col(sheet,6)
    sheet.print_area = f'F1:I{last_row}'

def set_print_area_priority_registry(sheet):
    last_row = find_last_row_in_col(sheet,2) # B column
    sheet.print_area = f'A1:I{last_row}'

def split_workbook(filename):
    '''
    This function splits the workbook into separate files and then zips them
    into a single file. It also deletes the files after they are added to the zip

    :param filename: The name of the file to be split
    '''

    initial_sheets = ['REESTR', 'СПР_ОБЪЕКТОВ', 'СПР_ПОДПИСАНТОВ']

    workbook = load_excel(f'saves/{filename}') # load a workbook
    filenames = [] # for the future filenames
    date_str = datetime.now().strftime('%d/%m/%Y_%H%M%S') # Format the current date

    for sheet in workbook.sheetnames:
        # if the sheet name is not in initial sheets
        if sheet not in initial_sheets:
            new_workbook = load_excel('template.xlsx')
            new_workbook.remove(new_workbook.active)
            new_workbook.add_sheet(workbook[sheet])
            new_workbook.save(f'saves/{sheet}_{date_str}.xlsx')
            filenames.append(f'saves/{sheet}_{date_str}.xlsx')

    filename.rstrip('.xlsx') # Remove .xlsx from the filename
    zipname = f'zips/{filename}.zip'
    with zipfile.ZipFile(zipname, 'a') as zipf:
        for file in filenames:
            zipf.write(file)

    # Delete the files after they are added to the zip
    for file in filenames:
        os.remove(file)

    return zipname
