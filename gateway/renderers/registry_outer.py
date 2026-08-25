"""Внешний реестр (для уполномоченной компании): один лист, позиции подряд,
фиксированный блок подписей.

Порт models/outer_registry.py. Исправления:
- путь к шаблону (в оригинале 'excel/templates/…' — endpoint был мёртв);
- неизвестный БИК больше не роняет реестр: банк подписывается самим БИКом;
- реквизиты собираются только когда счёт и БИК заполнены (guard стоял
  ПОСЛЕ обращения к словарю банков);
- «None к договору …» при пустом виде платежа;
- отсутствие позиций в колонке B не даёт TypeError;
- дата в шапке «Разрешение № от …» — дд.мм.гггг, а не полный timestamp
  с микросекундами.
"""
from datetime import datetime
from pathlib import Path

import openpyxl
import yaml
from openpyxl.styles import Alignment, Border, Font, Side

from .common import find_last_row_in_col, format_datetime, format_row

DATA_COLS = ["B", "C", "D", "E", "F", "G", "H", "I", "J"]
START_ROW = 13


def load_banks(path: Path) -> dict[str, str]:
    return yaml.safe_load(path.read_text(encoding="utf-8"))


def render_outer(entries: list[dict], template_path, banks: dict[str, str]):
    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook.active

    sheet["B7"] = f"Разрешение № от {datetime.today().strftime('%d.%m.%Y')}"

    row = START_ROW
    for i, item in enumerate(entries):
        format_row(sheet, row, DATA_COLS, height=100)
        sheet[f"B{row}"] = i + 1
        sheet[f"C{row}"] = item.get("organization", "")

        osnovanije = ""
        if item.get("zusaetzliches_vertrag"):
            osnovanije += f"Доп. соглашение №{item['zusaetzliches_vertrag']},"
        if item.get("prilozhenija"):
            osnovanije += f"{item['prilozhenija']},"

        payment_type = str(item.get("payment_type") or "")
        if payment_type:
            sheet[f"E{row}"] = payment_type
        if item.get("name_of_contract") and item.get("date_of_contract"):
            contract_str = (f"{item['name_of_contract']} от "
                            f"{format_datetime(item['date_of_contract'])}")
            osnovanije += contract_str
            suffix = f"к договору {contract_str}"
            sheet[f"E{row}"] = f"{payment_type} {suffix}".strip()
            sheet[f"F{row}"] = contract_str

        sheet[f"D{row}"] = osnovanije
        sheet[f"G{row}"] = float(item.get("contract_sum") or 0)
        sheet[f"H{row}"] = float(item.get("payment_sum") or 0)

        schet = item.get("schet_counter") or ""
        bik = item.get("BIK_counter") or ""
        if schet and bik:
            bank_name = banks.get(bik) or bik  # неизвестный БИК подписываем им самим
            sheet[f"I{row}"] = f"ИИК {schet}\nБИК {bik} в {bank_name}\n"
        row += 1

    _add_signatures(sheet)
    return workbook


def _add_signatures(sheet) -> None:
    """Фиксированный блок: уполномоченная компания, директор, исполнитель,
    вторая компания. Порт add_coordinators_outer как есть."""
    left14 = Alignment(horizontal="left")
    bold14 = Font(size=14, bold=True)
    thick_bottom = Border(bottom=Side(style="thick"))

    final_row = (find_last_row_in_col(sheet, 2) or START_ROW - 1) + 4
    sheet.cell(row=final_row, column=2, value='Уполномоченная компания: ТОО "New Line Project"')
    sheet.cell(row=final_row, column=2).alignment = left14
    sheet.cell(row=final_row, column=2).font = bold14
    sheet.cell(row=final_row, column=6, value="М.П.")

    final_row += 2
    sheet.cell(row=final_row, column=2, value="Директор:")
    sheet.cell(row=final_row, column=2).alignment = left14
    sheet.cell(row=final_row, column=2).font = bold14
    sheet.cell(row=final_row, column=5).border = thick_bottom
    sheet.cell(row=final_row + 1, column=5, value="(подпись)")
    sheet.cell(row=final_row, column=6, value="Бектемирова Ж.Ж")
    sheet.cell(row=final_row, column=6).font = Font(bold=True)

    final_row += 3
    sheet.cell(row=final_row, column=2, value="Исполнитель:")
    sheet.cell(row=final_row, column=2).alignment = left14
    sheet.cell(row=final_row, column=2).font = bold14
    sheet.cell(row=final_row, column=5).border = thick_bottom
    sheet.cell(row=final_row + 1, column=5, value="(подпись)")
    sheet.cell(row=final_row, column=6, value="Олжабаева Г.Т.")
    sheet.cell(row=final_row, column=6).font = Font(bold=True)
    for col in range(2, 8):
        sheet.cell(row=final_row + 2, column=col).border = thick_bottom

    final_row += 4
    sheet.cell(row=final_row, column=2, value='ТОО "Инжиниринговая компания "Лидер""')
    sheet.cell(row=final_row, column=2).alignment = Alignment(wrap_text=False)

    final_row += 2
    sheet.cell(row=final_row, column=2, value="Исполнитель:")
    sheet.cell(row=final_row, column=5).border = thick_bottom
    sheet.cell(row=final_row + 1, column=5, value="(подпись)")
    sheet.cell(row=final_row, column=6, value="Колоскова И.Б.")
