"""Общие помощники Excel-рендереров.

Порт utils/scripts.py с исправлениями:
- format_row / format_row_priority_registry схлопнуты в одну функцию
  (отличались только высотой строки и левым выравниванием F);
- удаление префиксов — removeprefix() вместо lstrip()/strip(), которые
  режут по набору символов и портят значения;
- create_concatenated_info не падает на None/числах из Doc-V.
"""
import re
from dataclasses import dataclass
from datetime import datetime
from typing import Optional

from openpyxl.styles import Alignment, Border, Font, Side

ARIAL_12 = Font(name="Arial", size=12)


def set_border(style: str) -> Border:
    side = Side(style=style)
    return Border(left=side, right=side, top=side, bottom=side)


def format_row(sheet, row_number: int, columns: list[str], *, height: int = 100,
               left_align: tuple[str, ...] = ()) -> None:
    """Arial 12, тонкая рамка, центрирование с переносом; height — высота строки.

    left_align — колонки, которым вместо центра нужно левое выравнивание
    (колонка сторон «Заявитель/Кому» во внутреннем реестре).
    """
    border = set_border("thin")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    sheet.row_dimensions[row_number].height = height
    for col in columns:
        cell = sheet[f"{col}{row_number}"]
        cell.font = ARIAL_12
        cell.border = border
        cell.alignment = left if col in left_align else center


def set_cell_properties(sheet, row, column, value, border=None, alignment=None, font=None):
    cell = sheet.cell(row=row, column=column, value=value)
    if border:
        cell.border = border
    if alignment:
        cell.alignment = alignment
    if font:
        cell.font = font


def hide_sheets(workbook, sheet_names) -> None:
    for name in sheet_names:
        workbook[name].sheet_state = "hidden"


def find_last_row_in_col(sheet, col_index: int) -> Optional[int]:
    for row in range(sheet.max_row, 0, -1):
        if sheet.cell(row=row, column=col_index).value:
            return row
    return None


def format_datetime(value) -> str:
    """ISO-строка -> дд/мм/гггг; всё, что не разбирается, возвращается как есть."""
    try:
        return datetime.fromisoformat(str(value)).strftime("%d/%m/%Y")
    except ValueError:
        return str(value)


def add_colontituls(sheet) -> None:
    sheet.oddFooter.left.text = "Группа компаний «Шар Құрылыс»"
    sheet.oddFooter.right.text = "Дата и время печати &D &T"


def set_print_area(sheet, *, anchor_col: str, anchor_col_index: int, area: str) -> None:
    """area — шаблон вида "F1:I{row}"; последняя строка ищется по anchor_col."""
    last = find_last_row_in_col(sheet, anchor_col_index) or 1
    sheet.print_area = area.format(row=last)


# ---------------------------------------------------------------------------
# Колонка «Назначение» внутреннего реестра
# ---------------------------------------------------------------------------

@dataclass
class FieldDef:
    key: str
    prefix: str = ""
    suffix: str = ""
    strip_prefix: str = ""
    placeholder: Optional[str] = None
    extract_numbers: bool = False


FIELD_CONFIG = [
    FieldDef(key="schet_na_oplatu", prefix="Счет на оплату №", strip_prefix="№"),
    FieldDef(key="esf", prefix="ЭСФ №", strip_prefix="№"),
    FieldDef(key="avr", prefix="Акт выполненных работ №", strip_prefix="№"),
    FieldDef(key="akt_sverki", prefix="Акт сверки ", strip_prefix="№"),
    FieldDef(key="sluzhebnaja_zapiska", prefix="Служебная записка "),
    FieldDef(key="avansovy_otchet", prefix="Авансовый отчет №", strip_prefix="№"),
    FieldDef(key="TRU"),
    FieldDef(key="letter", prefix="Письмо "),
    FieldDef(key="mediation", prefix="Медиация/Решение суда №", strip_prefix="№"),
    FieldDef(key="nakladnye", prefix="Накладные: ", extract_numbers=True),
    FieldDef(key="sogl_o_rastor", prefix="Согл. о расторжении №", strip_prefix="№",
             placeholder="placeholder", extract_numbers=True),
    FieldDef(key="prilozhenija", prefix="по приложению ", strip_prefix="Приложение "),
]


def _process_field(data_item: dict, field: FieldDef) -> Optional[str]:
    value = data_item.get(field.key) or ""
    value = str(value).strip()
    if not value:
        return None
    if field.placeholder and value == field.placeholder:
        return None
    if field.extract_numbers:
        match = re.search(r"\d+", value)
        if not match:
            return None
        value = value[match.start():].strip()
    if field.strip_prefix:
        value = value.removeprefix(field.strip_prefix).strip()
    if not value:
        return None
    return f"{field.prefix}{value}{field.suffix}"


def create_concatenated_info(data_item: dict) -> str:
    parts = []
    payment_type = str(data_item.get("payment_type") or "").strip()
    payment_objective = str(data_item.get("payment_objective") or "").strip()
    if payment_type and payment_type != payment_objective:
        parts.append(payment_type)
    if payment_objective:
        parts.append(payment_objective)

    for field in FIELD_CONFIG:
        result = _process_field(data_item, field)
        if result:
            parts.append(result)

    zusaetzliches_vertrag = data_item.get("zusaetzliches_vertrag") or ""
    if zusaetzliches_vertrag and zusaetzliches_vertrag != "placeholder":
        zv_text = str(zusaetzliches_vertrag).removeprefix("Доп. соглашение").strip()
        parts.append(f"ДС {zv_text}")
    else:
        if data_item.get("name_of_contract") and data_item.get("date_of_contract"):
            parts.append(f"Дог. {data_item['name_of_contract']}")

    payment_number = data_item.get("payment_number") or ""
    if payment_number:
        doctype = data_item.get("doctype") or ""
        parts.append(f"{doctype} №{payment_number}")

    return ", ".join(parts).rstrip(", ")
